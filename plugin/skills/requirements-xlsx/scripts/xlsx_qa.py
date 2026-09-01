#!/usr/bin/env python3
"""提出前の Excel を機械で検査する。元ファイルは変更しない。

    python3 xlsx_qa.py 機能一覧.xlsx --out-dir 90_qa/2026-09-01-1200

検査する内容
    1. 数式エラー（キャッシュされた計算結果と、式に残った #REF!）
    2. 外部ブックへのリンク
    3. 非表示のシート・行・列
    4. 名前定義（壊れているもの、どこからも使われていないもの）
    5. 入力規則
    6. シート間参照（存在しないシートを指しているもの）
    7. 同じ列の中での表示形式の揺れ

終了コード
    0 指摘なし / 1 警告あり / 2 提出できない指摘あり、またはファイルが読めない
"""
import argparse
import os
import re
import sys
import zipfile
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import qa_report  # noqa: E402

ERROR_VALUES = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#SPILL!"}
EXT_REF = re.compile(r"\[\d+\]|\[[^\]]+\.xls[xmb]?\]", re.I)
SHEET_REF = re.compile(r"(?:'([^']+)'|([^\s!()+\-*/,:=<>&]+))!\$?[A-Z]{1,3}\$?\d+")
OPEN_WORDS = re.compile(r"要確認|要検討|確認中|未確定|未決|未定|TBD|要相談|検討中|保留|要調整")


def add(f, sev, check, where, what, how=""):
    f.append({"重要度": sev, "検査": check, "場所": where, "内容": what, "対応": how})


def cell_ref(ws_title, coord):
    return "%s!%s" % (ws_title, coord)


def check_workbook(path):
    from openpyxl import load_workbook

    findings = []
    notes = []

    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)

    sheet_names = set(wb_f.sheetnames)
    used_names = set()
    cached_seen = False

    for ws in wb_f.worksheets:
        wsv = wb_v[ws.title]

        # --- 非表示シート
        if ws.sheet_state != "visible":
            add(findings, "警告", "非表示シート", ws.title,
                "シートが %s になっている" % ws.sheet_state,
                "提出前に、意図した非表示か確認する")

        # --- 非表示の行・列
        hidden_rows = [str(i) for i, d in ws.row_dimensions.items() if d.hidden]
        if hidden_rows:
            add(findings, "警告", "非表示行", ws.title,
                "非表示の行 %d 件（%s）" % (len(hidden_rows), ",".join(hidden_rows[:10])),
                "隠したまま提出していないか確認する")
        hidden_cols = [c for c, d in ws.column_dimensions.items() if d.hidden]
        if hidden_cols:
            add(findings, "警告", "非表示列", ws.title,
                "非表示の列 %d 件（%s）" % (len(hidden_cols), ",".join(hidden_cols[:10])),
                "隠したまま提出していないか確認する")

        # --- 入力規則
        try:
            dvs = list(ws.data_validations.dataValidation)
        except Exception:
            dvs = []
        for dv in dvs:
            add(findings, "情報", "入力規則", "%s!%s" % (ws.title, dv.sqref),
                "%s（%s）" % (dv.type, (dv.formula1 or "")[:60]),
                "凡例シートの選択肢と食い違っていないか確認する")

        fmt_by_col = defaultdict(set)

        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue

                # --- 数式
                if isinstance(v, str) and v.startswith("="):
                    if "#REF!" in v:
                        add(findings, "要確認", "数式エラー", cell_ref(ws.title, c.coordinate),
                            "式に #REF! が残っている: %s" % v[:80],
                            "参照先を直すか、式を消す")
                    if EXT_REF.search(v):
                        add(findings, "要確認", "外部リンク", cell_ref(ws.title, c.coordinate),
                            "外部ブックを参照している: %s" % v[:80],
                            "リンク先を開く前に発信元へ確認する。値に置き換えて提出する")
                    for m in SHEET_REF.finditer(v):
                        name = m.group(1) or m.group(2)
                        if name in {"TRUE", "FALSE"}:
                            continue
                        used_names.add(name)
                        if name not in sheet_names and not EXT_REF.search(v):
                            add(findings, "警告", "シート間参照",
                                cell_ref(ws.title, c.coordinate),
                                "存在しないシート「%s」を参照している" % name,
                                "シート名の変更か削除の取り残しを確認する")

                # --- 未決セル（コメントの有無）
                if isinstance(v, str) and OPEN_WORDS.search(v):
                    if c.comment is None:
                        add(findings, "警告", "未決セル", cell_ref(ws.title, c.coordinate),
                            "未決だがコメントが付いていない: %s" % v[:40],
                            "誰が・いつまでに・何を決めるかをコメントに書く")

                # --- 表示形式の揺れ
                if c.number_format and c.number_format != "General":
                    fmt_by_col[c.column_letter].add(c.number_format)

                # --- キャッシュされた計算結果のエラー（数式セルだけを見る）
                if isinstance(v, str) and v.startswith("="):
                    cv = wsv[c.coordinate].value
                    if cv is not None:
                        cached_seen = True
                    if isinstance(cv, str) and cv.strip() in ERROR_VALUES:
                        add(findings, "要確認", "数式エラー", cell_ref(ws.title, c.coordinate),
                            "計算結果が %s になっている" % cv.strip(),
                            "式か参照先を直す")

        for col, fmts in fmt_by_col.items():
            if len(fmts) > 1:
                add(findings, "警告", "表示形式の揺れ", "%s!%s列" % (ws.title, col),
                    "同じ列に %d 種類の表示形式がある（%s）"
                    % (len(fmts), " / ".join(sorted(fmts)[:4])),
                    "単位・日付・小数点の桁をそろえる")

    # --- 名前定義
    try:
        defined = dict(wb_f.defined_names)
    except Exception:
        defined = {}
    for name, dn in defined.items():
        dest = getattr(dn, "attr_text", "") or ""
        if "#REF!" in dest:
            add(findings, "要確認", "名前定義", name,
                "参照先が #REF! になっている", "名前定義を消すか参照先を直す")
        elif name not in used_names:
            add(findings, "警告", "名前定義", name,
                "どの式からも使われていない（参照先 %s）" % dest[:60],
                "使っていないなら提出前に消す")

    # --- 外部リンクの部品そのもの
    try:
        with zipfile.ZipFile(path) as z:
            ext = [n for n in z.namelist() if n.startswith("xl/externalLinks/") and n.endswith(".xml")]
    except Exception:
        ext = []
    if ext:
        add(findings, "要確認", "外部リンク", "ブック全体",
            "外部リンクの定義が %d 件ある" % len(ext),
            "リンクを解除して値にするか、発信元へ確認する")

    if not cached_seen:
        notes.append("このブックには計算結果が保存されていないため、"
                     "0 除算などの実行時エラーは検出できない。"
                     "Excel で一度開いて保存したファイルなら検出できる。")
    notes.append("グラフと元データの値が合っているかは、この検査では判定しない。目視で確認する。")

    return findings, notes


def main():
    ap = argparse.ArgumentParser(description="Excel の機械検査（元ファイルは変更しない）")
    ap.add_argument("file", help="検査する .xlsx")
    ap.add_argument("--out-dir", default=".", help="結果の出力先")
    ap.add_argument("--stem", help="出力ファイル名（既定: xlsx-<ファイル名>）")
    args = ap.parse_args()

    if not os.path.isfile(args.file):
        print("ファイルがありません: %s" % args.file, file=sys.stderr)
        return 2

    versions = qa_report.tool_versions(["openpyxl"])
    try:
        findings, notes = check_workbook(args.file)
    except Exception as e:
        print("読めませんでした: %s: %s" % (type(e).__name__, e), file=sys.stderr)
        return 2

    report = qa_report.build("Excel", args.file, findings, versions, notes)
    stem = args.stem or ("xlsx-" + os.path.splitext(os.path.basename(args.file))[0])
    jp, mp = qa_report.write(report, args.out_dir, stem)
    c = report["件数"]
    print("Excel  %s  要確認 %d / 警告 %d / 情報 %d  → %s"
          % (os.path.basename(args.file), c["要確認"], c["警告"], c["情報"], mp))
    return qa_report.exit_code(report)


if __name__ == "__main__":
    sys.exit(main())
